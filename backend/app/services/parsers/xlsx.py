import io
import openpyxl

def is_tabular_sheet(sheet) -> bool:
    """
    Heuristic to determine if a sheet contains tabular financial/contractual data
    vs metadata, charts, or empty pages.
    """
    # If the sheet is empty or has very small dimensions
    if sheet.max_row <= 1 or sheet.max_column <= 1:
        return False
        
    filled_cell_count = 0
    row_cell_counts = []
    
    # Analyze the first 50 rows (for speed/safety)
    for row in sheet.iter_rows(max_row=50, values_only=True):
        non_empty_in_row = sum(1 for val in row if val is not None)
        if non_empty_in_row > 0:
            filled_cell_count += non_empty_in_row
            row_cell_counts.append(non_empty_in_row)
            
    if filled_cell_count < 6:
        # Too little data, likely metadata or empty
        return False
        
    # If the average filled cells per active row is low, or it's mostly single key-value entries
    avg_filled_per_row = sum(row_cell_counts) / len(row_cell_counts) if row_cell_counts else 0
    max_filled_in_row = max(row_cell_counts) if row_cell_counts else 0
    
    # Tabular sheets typically have at least one row with 3+ fields, and average > 2 fields per row
    if max_filled_in_row >= 3 and avg_filled_per_row >= 2.0:
        return True
        
    return False

def parse_xlsx(file_bytes: bytes) -> tuple[str, float, str]:
    """
    Parses Excel sheets, identifies tabular data vs metadata sheets, and processes them separately.
    Returns:
        extracted_text (str)
        confidence (float): 100.0
        warning_message (str)
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    output_parts = []
    skipped_sheets = []
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        if is_tabular_sheet(sheet):
            output_parts.append(f"### Sheet: {sheet_name} (Tabular Data)")
            
            # Format sheet content as markdown table
            table_rows = []
            for row in sheet.iter_rows(values_only=True):
                # Skip empty rows
                if all(val is None for val in row):
                    continue
                    
                # Format cell values
                row_str = []
                for val in row:
                    if val is None:
                        row_str.append("")
                    elif isinstance(val, float):
                        # Clean currency or percentage formatting
                        row_str.append(f"{val:.2f}")
                    else:
                        row_str.append(str(val).strip().replace("\n", " "))
                
                # Truncate empty cells from right
                while row_str and row_str[-1] == "":
                    row_str.pop()
                    
                if row_str:
                    table_rows.append(row_str)
            
            if not table_rows:
                continue
                
            # Construct markdown table structure
            max_cols = max(len(r) for r in table_rows)
            formatted_table = []
            
            for idx, r in enumerate(table_rows):
                # Pad row to max columns
                padded_r = r + [""] * (max_cols - len(r))
                formatted_table.append("| " + " | ".join(padded_r) + " |")
                
                # Add divider after the first row (assumed header)
                if idx == 0:
                    formatted_table.append("|" + "|".join(["---"] * max_cols) + "|")
                    
            output_parts.append("\n".join(formatted_table) + "\n")
        else:
            # Metadata or minor sheets
            skipped_sheets.append(sheet_name)
            output_parts.append(f"### Sheet: {sheet_name} (Metadata/Reference)")
            
            # Process as simple key-value lists
            for row in sheet.iter_rows(values_only=True):
                # Clean row values
                non_empty = [str(val).strip() for val in row if val is not None]
                if len(non_empty) == 2:
                    output_parts.append(f"- **{non_empty[0]}**: {non_empty[1]}")
                elif len(non_empty) == 1:
                    output_parts.append(f"- {non_empty[0]}")
                elif len(non_empty) > 2:
                    output_parts.append(f"- " + ", ".join(non_empty))
            output_parts.append("")
            
    full_text = "\n\n".join(output_parts).strip()
    warning = ""
    if skipped_sheets:
        warning = f"Filtered {len(skipped_sheets)} metadata or non-tabular sheets: {', '.join(skipped_sheets)}."
        
    return full_text, 100.0, warning
